import os
import hashlib
import io
import logging
import secrets
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from fastapi import APIRouter, Cookie, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app import database, models, schemas

router = APIRouter()
logger = logging.getLogger("admin")

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin_username")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin_password")

if ADMIN_PASSWORD == "admin_password":
    logger.warning(
        "ADMIN_PASSWORD is not set — running with the default password. "
        "Set ADMIN_USERNAME and ADMIN_PASSWORD env vars before running with a real class."
    )

ADMIN_COOKIE_NAME = "admin_token"
ADMIN_COOKIE_MAX_AGE = 60 * 60 * 24  # 24 hours


def _compute_admin_token() -> str:
    """
    Deterministic token derived from the hardcoded credentials. This keeps
    admin auth stateless (no extra DB table needed, and it survives server
    restarts/`--reload`) while avoiding putting the raw password directly
    in the browser's cookie.
    """
    raw = f"{ADMIN_USERNAME}:{ADMIN_PASSWORD}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


_EXPECTED_ADMIN_TOKEN = _compute_admin_token()


def get_current_admin(admin_token: Optional[str] = Cookie(default=None)) -> None:
    """Dependency that guards every other endpoint in this router."""
    if not admin_token or not secrets.compare_digest(admin_token, _EXPECTED_ADMIN_TOKEN):
        raise HTTPException(status_code=401, detail="دسترسی مدیریت ندارید، وارد شوید")


# ==========================================
# Authentication
# ==========================================

@router.post("/login")
def admin_login(credentials: schemas.AdminLogin, response: Response):
    valid = secrets.compare_digest(
        credentials.username, ADMIN_USERNAME
    ) and secrets.compare_digest(credentials.password, ADMIN_PASSWORD)

    if not valid:
        logger.warning("Failed admin login attempt (username=%s)", credentials.username)
        raise HTTPException(status_code=401, detail="نام کاربری یا رمز عبور اشتباه است")

    response.set_cookie(
        key=ADMIN_COOKIE_NAME,
        value=_EXPECTED_ADMIN_TOKEN,
        max_age=ADMIN_COOKIE_MAX_AGE,
        httponly=True,
        samesite="lax",
    )
    logger.info("Admin logged in successfully")
    return {"message": "ورود موفقیت‌آمیز بود"}


# ==========================================
# Lobby waiting list
# ==========================================

@router.get("/waiting-list", response_model=schemas.StudentAdminListOut)
def get_waiting_list(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    """Lets the teacher see, before starting the game, who has registered and is waiting."""
    students = (
        db.query(models.Student).order_by(models.Student.created_at.asc()).all()
    )
    return schemas.StudentAdminListOut(count=len(students), students=students)

# ==========================================
# Game state control
# ==========================================

@router.post("/game/start", response_model=schemas.GameStatusOut)
def start_game(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    game_state = db.query(models.GameState).first()
    if game_state is None:
        raise HTTPException(status_code=500, detail="game_state تعریف نشده است")

    settings = db.query(models.GameSettings).first()
    if settings is not None:
        settings.starting_time = datetime.now(timezone.utc)

    game_state.status = "playing"
    db.commit()
    logger.info("Admin started the game")
    return game_state


@router.post("/game/stop", response_model=schemas.GameStatusOut)
def stop_game(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    game_state = db.query(models.GameState).first()
    if game_state is None:
        raise HTTPException(status_code=500, detail="game_state تعریف نشده است")

    game_state.status = "finished"
    db.commit()
    logger.info("Admin stopped the game")
    return game_state


@router.post("/game/reset")
def reset_game(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    """
    Clears all students AND their logged answers (student_answers references
    students, so answers are removed first to avoid orphaned rows), and
    returns the game to 'waiting'.
    """
    removed_answers = db.query(models.StudentAnswer).delete()
    removed_students = db.query(models.Student).delete()

    game_state = db.query(models.GameState).first()
    if game_state is not None:
        game_state.status = "waiting"

    settings = db.query(models.GameSettings).first()
    if settings is not None:
        settings.starting_time = None

    db.commit()
    logger.info(
        "Admin reset the game: removed %s students and %s answers",
        removed_students,
        removed_answers,
    )
    return {"status": "waiting", "removed_students": removed_students}


# ==========================================
# Game timing settings
# ==========================================

@router.get("/settings", response_model=schemas.GameSettingsOut)
def get_settings(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    settings = db.query(models.GameSettings).first()
    if settings is None:
        raise HTTPException(status_code=500, detail="تنظیمات بازی تعریف نشده است")
    return settings


@router.put("/settings", response_model=schemas.GameSettingsOut)
def update_settings(
    payload: schemas.GameSettingsUpdate,
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    """Both fields are optional — send just the one the teacher changed."""
    settings = db.query(models.GameSettings).first()
    if settings is None:
        raise HTTPException(status_code=500, detail="تنظیمات بازی تعریف نشده است")

    if payload.game_time is not None:
        settings.game_time = payload.game_time
    if payload.guid_time is not None:
        settings.guid_time = payload.guid_time

    db.commit()
    logger.info(
        "Admin updated settings: game_time=%s guid_time=%s",
        settings.game_time,
        settings.guid_time,
    )
    return settings


# ==========================================
# Results export (Excel)
# ==========================================

@router.get("/results/export")
def export_results_excel(
    db: Session = Depends(database.get_db),
    _: None = Depends(get_current_admin),
):
    """
    Builds an .xlsx with two sheets: final rankings, and every individual
    answer broken down by question — the second sheet is for the kind of
    item-level analysis (which questions the class struggled with) the
    project's own architecture notes mentioned wanting later.
    """
    students = (
        db.query(models.Student)
        .order_by(models.Student.score.desc(), models.Student.created_at.asc())
        .all()
    )

    workbook = openpyxl.Workbook()
    header_font = Font(bold=True)

    results_sheet = workbook.active
    results_sheet.title = "نتایج"
    results_sheet.sheet_view.rightToLeft = True
    results_sheet.append(["رتبه", "نام دانش‌آموز", "امتیاز", "زمان ورود به لابی", "آی‌پی", "دستگاه"])
    for cell in results_sheet[1]:
        cell.font = header_font

    for rank, student in enumerate(students, start=1):
        results_sheet.append([
            rank, student.name, student.score,
            student.created_at.strftime("%Y-%m-%d %H:%M:%S") if student.created_at else "",
            student.ip_address or "", student.device_info or "",
        ])

    detail_sheet = workbook.create_sheet(title="پاسخ به تفکیک سوال")
    detail_sheet.sheet_view.rightToLeft = True
    detail_sheet.append(["نام دانش‌آموز", "شماره سوال", "گزینه انتخابی", "نتیجه"])
    for cell in detail_sheet[1]:
        cell.font = header_font

    answers = (
        db.query(models.StudentAnswer)
        .join(models.Student)
        .order_by(models.Student.name, models.StudentAnswer.question_id)
        .all()
    )
    for ans in answers:
        detail_sheet.append([
            ans.student.name,
            ans.question_id,
            ans.selected_option,
            "درست" if ans.is_correct else "غلط",
        ])

    for sheet in (results_sheet, detail_sheet):
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 22

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    logger.info(
        "Admin exported results: %s students, %s logged answers",
        len(students),
        len(answers),
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quiz-results.xlsx"},
    )