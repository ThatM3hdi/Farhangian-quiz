import io

import openpyxl

from app.routers.admin import ADMIN_PASSWORD, ADMIN_USERNAME


def _login_as_admin(client):
    return client.post(
        "/api/admin/login",
        json={"username": ADMIN_USERNAME, "password": ADMIN_PASSWORD},
    )


def _register(client, name="دانش‌آموز تست"):
    return client.post("/api/lobby/register", json={"name": name})


def test_admin_login_success_sets_cookie(client):
    response = _login_as_admin(client)

    assert response.status_code == 200
    assert "admin_token" in response.cookies


def test_admin_login_wrong_password_rejected(client):
    response = client.post(
        "/api/admin/login",
        json={"username": ADMIN_USERNAME, "password": "wrong-password"},
    )
    assert response.status_code == 401


def test_admin_endpoint_requires_login(client):
    response = client.get("/api/admin/waiting-list")
    assert response.status_code == 401


def test_waiting_list_shows_registered_students(client):
    _login_as_admin(client)
    _register(client, "دانش‌آموز یک")
    _register(client, "دانش‌آموز دو")

    response = client.get("/api/admin/waiting-list")

    assert response.status_code == 200
    data = response.json()
    assert data["count"] == 2
    names = [s["name"] for s in data["students"]]
    assert "دانش‌آموز یک" in names
    assert "دانش‌آموز دو" in names


def test_start_stop_reset_game_flow(client):
    _login_as_admin(client)

    start_response = client.post("/api/admin/game/start")
    assert start_response.status_code == 200
    assert start_response.json()["status"] == "playing"
    assert client.get("/api/lobby/status").json()["status"] == "playing"

    stop_response = client.post("/api/admin/game/stop")
    assert stop_response.json()["status"] == "finished"

    _register(client, "دانش‌آموز جامانده")
    reset_response = client.post("/api/admin/game/reset")

    assert reset_response.json()["status"] == "waiting"
    assert reset_response.json()["removed_students"] == 1
    assert client.get("/api/admin/waiting-list").json()["count"] == 0


def test_get_settings_default_values(client):
    _login_as_admin(client)
    response = client.get("/api/admin/settings")

    assert response.status_code == 200
    data = response.json()
    assert data["game_time"] == 300
    assert data["guid_time"] == 20


def test_update_settings_partial(client):
    _login_as_admin(client)

    response = client.put("/api/admin/settings", json={"game_time": 400})

    assert response.status_code == 200
    data = response.json()
    assert data["game_time"] == 400
    assert data["guid_time"] == 20  # دست‌نخورده باقی می‌ماند


def test_update_settings_rejects_non_positive_values(client):
    _login_as_admin(client)

    response = client.put("/api/admin/settings", json={"game_time": 0})
    assert response.status_code == 422


def test_export_results_returns_valid_xlsx(playing_client):
    _login_as_admin(playing_client)
    _register(playing_client, "دانش‌آموز خروجی")
    playing_client.post(
        "/api/game/answer", json={"question_id": 1, "selected_option": 1}
    )

    response = playing_client.get("/api/admin/results/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "نتایج" in workbook.sheetnames
    assert "پاسخ به تفکیک سوال" in workbook.sheetnames

    results_header = [cell.value for cell in workbook["نتایج"][1]]
    assert "امتیاز" in results_header
